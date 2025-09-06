"""
Infrastructure adapters for data export.

This module contains concrete implementations of data export services
that handle exporting portfolio data to various formats.
"""

import csv
import os
from abc import ABC, abstractmethod
from typing import List, Dict, Any
import logging

from ..domain.entities import PortfolioPosition, Money, ProfitLoss
from ..application.use_cases import DataExportService

logger = logging.getLogger(__name__)


class BaseFileExporter(DataExportService, ABC):
    """Base class for file-based exporters."""
    
    def __init__(self, output_directory: str = "."):
        """
        Initialize base exporter.
        
        Args:
            output_directory: Directory where files will be saved
        """
        self._output_directory = output_directory
        os.makedirs(output_directory, exist_ok=True)
    
    def export_portfolio_data(self, positions: List[PortfolioPosition]) -> None:
        """Export portfolio data to files."""
        if not positions:
            logger.warning("No positions to export")
            return
        
        # Convert positions to exportable format
        portfolio_data = self._convert_positions_to_records(positions)
        yearly_data = self._aggregate_yearly_data(positions)
        
        # Export both datasets
        portfolio_file = self._export_portfolio_history(portfolio_data)
        yearly_file = self._export_yearly_summary(yearly_data)
        
        # Log export results
        self._log_export_results(portfolio_file, yearly_file, len(positions))
    
    def _convert_positions_to_records(self, positions: List[PortfolioPosition]) -> List[Dict[str, Any]]:
        """Convert positions to dictionary records for export."""
        records = []
        
        for position in positions:
            # Determine operation description
            operation_description = "N/A"
            operation_quantity = 0
            
            if position.operation_type and position.operation_quantity:
                operation_quantity = position.operation_quantity.value
                if position.operation_type == 'vesting':
                    operation_description = f"Vesting (+{operation_quantity})"
                elif position.operation_type == 'trade':
                    operation_description = f"Trade (-{operation_quantity})"
            
            record = {
                'Date': position.last_updated.strftime('%d/%m/%Y'),
                'Operation Description': operation_description,
                'Operation Quantity': operation_quantity,
                'Final Quantity': position.quantity.value,
                'Total Cost USD': Money(position.total_cost_usd.amount, 'USD'),
                'Average Price USD': Money(position.average_price_usd.amount, 'USD'),
                'Total Cost BRL': Money(position.total_cost_brl.amount, 'BRL'),
                'Average Price BRL': Money(position.average_price_brl.amount, 'BRL'),
                'Gross Profit BRL': ProfitLoss(position.gross_profit_brl.amount, 'BRL')
            }
            records.append(record)
        
        return records
    
    def _aggregate_yearly_data(self, positions: List[PortfolioPosition]) -> List[Dict[str, Any]]:
        """Aggregate positions into yearly summaries."""
        yearly_data = {}
        
        for position in positions:
            year = position.last_updated.year
            
            # Keep only the last position of each year
            # Use >= to handle multiple positions on the same date (keep the latest one processed)
            if year not in yearly_data or position.last_updated >= yearly_data[year]['last_date']:
                yearly_data[year] = {
                    'year': year,
                    'final_position': position,
                    'last_date': position.last_updated
                }
        
        # Convert to records
        records = []
        for year_data in yearly_data.values():
            position = year_data['final_position']
            record = {
                'Year': year_data['year'],
                'Final Quantity': position.quantity.value,
                'Total Cost USD': Money(position.total_cost_usd.amount, 'USD'),
                'Average Price USD': Money(position.average_price_usd.amount, 'USD'),
                'Total Cost BRL': Money(position.total_cost_brl.amount, 'BRL'),
                'Average Price BRL': Money(position.average_price_brl.amount, 'BRL'),
                'Gross Profit BRL': ProfitLoss(position.gross_profit_brl.amount, 'BRL')
            }
            records.append(record)
        
        # Sort by year
        records.sort(key=lambda x: x['Year'])
        return records
    
    @abstractmethod
    def _export_portfolio_history(self, data: List[Dict[str, Any]]) -> str:
        """Export portfolio history data and return filename."""
        pass
    
    @abstractmethod
    def _export_yearly_summary(self, data: List[Dict[str, Any]]) -> str:
        """Export yearly summary data and return filename."""
        pass
    
    def _log_export_results(self, portfolio_file: str, yearly_file: str, position_count: int) -> None:
        """Log export results."""
        print(f"\nData exported successfully:")
        print(f"  - Portfolio History: {os.path.abspath(portfolio_file)} ({position_count} records)")
        print(f"  - Yearly Summary: {os.path.abspath(yearly_file)}")


class CSVExporter(BaseFileExporter):
    """CSV file exporter."""
    
    def _export_portfolio_history(self, data: List[Dict[str, Any]]) -> str:
        """Export portfolio history to CSV."""
        filename = os.path.join(self._output_directory, "portfolio_history.csv")
        
        if not data:
            # Create empty file with headers
            with open(filename, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['Date', 'Operation Description', 'Operation Quantity', 'Final Quantity', 
                               'Total Cost USD', 'Average Price USD', 'Total Cost BRL', 'Average Price BRL', 'Gross Profit BRL'])
            return filename
        
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
        
        return filename
    
    def _export_yearly_summary(self, data: List[Dict[str, Any]]) -> str:
        """Export yearly summary to CSV."""
        filename = os.path.join(self._output_directory, "yearly_summary.csv")
        
        if not data:
            # Create empty file with headers
            with open(filename, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['Year', 'Final Quantity', 'Total Cost USD', 'Average Price USD',
                               'Total Cost BRL', 'Average Price BRL', 'Gross Profit BRL'])
            return filename
        
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
        
        return filename


class XLSXExporter(BaseFileExporter):
    """Excel file exporter."""
    
    def __init__(self, output_directory: str = "."):
        """Initialize XLSX exporter."""
        super().__init__(output_directory)
        try:
            import openpyxl
            self._openpyxl = openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl package is required for XLSX export. "
                "Install with: pip install openpyxl"
            )
    
    def _export_portfolio_history(self, data: List[Dict[str, Any]]) -> str:
        """Export portfolio history to XLSX."""
        filename = os.path.join(self._output_directory, "portfolio_history.xlsx")
        
        workbook = self._openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Portfolio History"
        
        if data:
            # Write headers
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                worksheet.cell(row=1, column=col, value=header)
            
            # Write data
            for row, record in enumerate(data, 2):
                for col, header in enumerate(headers, 1):
                    value = record[header]
                    # Convert Money and ProfitLoss objects to float for Excel compatibility
                    if hasattr(value, 'amount'):
                        value = float(value.amount)
                    worksheet.cell(row=row, column=col, value=value)
        
        workbook.save(filename)
        return filename
    
    def _export_yearly_summary(self, data: List[Dict[str, Any]]) -> str:
        """Export yearly summary to XLSX."""
        filename = os.path.join(self._output_directory, "yearly_summary.xlsx")
        
        workbook = self._openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Yearly Summary"
        
        if data:
            # Write headers
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                worksheet.cell(row=1, column=col, value=header)
            
            # Write data
            for row, record in enumerate(data, 2):
                for col, header in enumerate(headers, 1):
                    value = record[header]
                    # Convert Money and ProfitLoss objects to float for Excel compatibility
                    if hasattr(value, 'amount'):
                        value = float(value.amount)
                    worksheet.cell(row=row, column=col, value=value)
        
        workbook.save(filename)
        return filename


class NoOpExporter(DataExportService):
    """
    No-operation exporter that doesn't export anything.
    
    This is useful when export functionality is not needed.
    """
    
    def export_portfolio_data(self, positions: List[PortfolioPosition]) -> None:
        """No-op export (does nothing)."""
        logger.info(f"No-op export called for {len(positions)} positions")
        pass


class ExporterFactory:
    """Factory for creating appropriate exporters."""
    
    _exporters = {
        'csv': CSVExporter,
        'xlsx': XLSXExporter,
        'none': NoOpExporter
    }
    
    @classmethod
    def create_exporter(cls, export_format: str, output_directory: str = ".") -> DataExportService:
        """
        Create an exporter for the specified format.
        
        Args:
            export_format: Format to export ('csv', 'xlsx', or 'none')
            output_directory: Directory where files will be saved
            
        Returns:
            Appropriate exporter instance
            
        Raises:
            ValueError: If export format is not supported
        """
        if export_format not in cls._exporters:
            raise ValueError(f"Unsupported export format: {export_format}")
        
        exporter_class = cls._exporters[export_format]
        
        if export_format == 'none':
            return exporter_class()
        else:
            return exporter_class(output_directory)
