"""
Tests for export functionality with year-end reset.

This module contains tests to verify that the export functionality
correctly handles and displays year-end profit resets in CSV and XLSX formats.
"""

import pytest
import tempfile
import os
import csv
from datetime import datetime
from decimal import Decimal

from src.domain.entities import Money, StockQuantity, PortfolioPosition
from src.domain.operations import VestingOperation, TradeOperation
from src.domain.services import PortfolioCalculationService
from src.infrastructure.exchange_rate import MockExchangeRateService
from src.infrastructure.export import CSVExporter, XLSXExporter, ExporterFactory


class TestExportYearReset:
    """Test export functionality with year-end resets."""
    
    @pytest.fixture
    def multi_year_positions(self):
        """Create positions spanning multiple years with resets."""
        exchange_service = MockExchangeRateService()
        calc_service = PortfolioCalculationService(exchange_service)
        
        operations = [
            # 2022
            VestingOperation(
                date=datetime(2022, 2, 15),
                quantity=StockQuantity(100),
                price_per_share_usd=Money(Decimal('8.00'), 'USD')
            ),
            TradeOperation(
                date=datetime(2022, 8, 20),
                quantity=StockQuantity(40),
                price_per_share_usd=Money(Decimal('12.00'), 'USD')
            ),
            TradeOperation(
                date=datetime(2022, 11, 10),
                quantity=StockQuantity(20),
                price_per_share_usd=Money(Decimal('15.00'), 'USD')
            ),
            
            # 2023 - Reset should happen
            VestingOperation(
                date=datetime(2023, 1, 25),
                quantity=StockQuantity(60),
                price_per_share_usd=Money(Decimal('10.00'), 'USD')
            ),
            TradeOperation(
                date=datetime(2023, 6, 15),
                quantity=StockQuantity(30),
                price_per_share_usd=Money(Decimal('18.00'), 'USD')
            ),
            
            # 2024 - Another reset
            VestingOperation(
                date=datetime(2024, 3, 10),
                quantity=StockQuantity(50),
                price_per_share_usd=Money(Decimal('14.00'), 'USD')
            ),
            TradeOperation(
                date=datetime(2024, 9, 5),
                quantity=StockQuantity(25),
                price_per_share_usd=Money(Decimal('20.00'), 'USD')
            )
        ]
        
        return calc_service.execute_operations(operations)
    
    def test_csv_export_shows_year_resets(self, multi_year_positions):
        """Test that CSV export correctly shows year-end resets."""
        with tempfile.TemporaryDirectory() as temp_dir:
            exporter = CSVExporter(temp_dir)
            exporter.export_portfolio_data(multi_year_positions)
            
            # Read portfolio history CSV
            portfolio_file = os.path.join(temp_dir, 'portfolio_history.csv')
            assert os.path.exists(portfolio_file)
            
            with open(portfolio_file, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                rows = list(reader)
            
            # Should have 7 rows (7 operations)
            assert len(rows) == 7
            
            # Verify headers include new fields
            expected_headers = {
                'Date', 'Operation Description', 'Operation Quantity', 
                'Final Quantity', 'Total Cost USD', 'Average Price USD',
                'Total Cost BRL', 'Average Price BRL', 'Gross Profit BRL'
            }
            assert set(rows[0].keys()) == expected_headers
            
            # Parse and verify data
            profits = []
            years = []
            operations = []
            
            for row in rows:
                date_str = row['Date']
                year = int(date_str.split('/')[-1])
                profit = float(row['Gross Profit BRL'])
                operation = row['Operation Description']
                
                years.append(year)
                profits.append(profit)
                operations.append(operation)
            
            # Verify year progression
            assert years == [2022, 2022, 2022, 2023, 2023, 2024, 2024]
            
            # Verify profit patterns with resets
            # 2022: 0 -> profit -> more profit
            assert profits[0] == 0.0  # First vesting
            assert profits[1] > 0.0   # First trade
            assert profits[2] > profits[1]  # Second trade (accumulated)
            
            # 2023: Reset to 0 -> profit
            assert profits[3] == 0.0  # Reset after vesting
            assert profits[4] > 0.0   # Trade profit (fresh start)
            
            # 2024: Reset to 0 -> profit
            assert profits[5] == 0.0  # Reset after vesting
            assert profits[6] > 0.0   # Trade profit (fresh start)
            
            # Verify profits are independent between years
            assert profits[2] != profits[4]  # 2022 final != 2023 final
            assert profits[4] != profits[6]  # 2023 final != 2024 final
            
            # Verify operation descriptions
            assert 'Vesting (+100)' in operations[0]
            assert 'Trade (-40)' in operations[1]
            assert 'Trade (-20)' in operations[2]
            assert 'Vesting (+60)' in operations[3]
            assert 'Trade (-30)' in operations[4]
            assert 'Vesting (+50)' in operations[5]
            assert 'Trade (-25)' in operations[6]
    
    def test_yearly_summary_csv_shows_final_year_profits(self, multi_year_positions):
        """Test that yearly summary shows only final year profits (no carryover)."""
        with tempfile.TemporaryDirectory() as temp_dir:
            exporter = CSVExporter(temp_dir)
            exporter.export_portfolio_data(multi_year_positions)
            
            # Read yearly summary CSV
            yearly_file = os.path.join(temp_dir, 'yearly_summary.csv')
            assert os.path.exists(yearly_file)
            
            with open(yearly_file, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                rows = list(reader)
            
            # Should have 3 rows (3 years: 2022, 2023, 2024)
            assert len(rows) == 3
            
            # Verify headers
            expected_headers = {
                'Year', 'Final Quantity', 'Total Cost USD', 'Average Price USD',
                'Total Cost BRL', 'Average Price BRL', 'Gross Profit BRL'
            }
            assert set(rows[0].keys()) == expected_headers
            
            # Parse yearly data
            yearly_data = {}
            for row in rows:
                year = int(row['Year'])
                profit = float(row['Gross Profit BRL'])
                yearly_data[year] = profit
            
            # Verify years
            assert set(yearly_data.keys()) == {2022, 2023, 2024}
            
            # Each year should have independent profits
            assert yearly_data[2022] > 0  # Had trades
            assert yearly_data[2023] > 0  # Had trade
            assert yearly_data[2024] > 0  # Had trade
            
            # Profits should be different (independent calculations)
            profits = list(yearly_data.values())
            assert len(set(profits)) == len(profits)  # All unique
    
    def test_xlsx_export_with_year_resets(self, multi_year_positions):
        """Test XLSX export with year-end resets."""
        with tempfile.TemporaryDirectory() as temp_dir:
            try:
                exporter = XLSXExporter(temp_dir)
                exporter.export_portfolio_data(multi_year_positions)
                
                # Verify files were created
                portfolio_file = os.path.join(temp_dir, 'portfolio_history.xlsx')
                yearly_file = os.path.join(temp_dir, 'yearly_summary.xlsx')
                
                assert os.path.exists(portfolio_file)
                assert os.path.exists(yearly_file)
                
                # Verify files are not empty
                assert os.path.getsize(portfolio_file) > 0
                assert os.path.getsize(yearly_file) > 0
                
                # Basic validation using openpyxl
                import openpyxl
                
                # Check portfolio history workbook
                wb_portfolio = openpyxl.load_workbook(portfolio_file)
                ws_portfolio = wb_portfolio.active
                assert ws_portfolio.title == "Portfolio History"
                
                # Should have header + 7 data rows
                assert ws_portfolio.max_row == 8  # Header + 7 operations
                assert ws_portfolio.max_column == 9  # 9 columns
                
                # Check yearly summary workbook
                wb_yearly = openpyxl.load_workbook(yearly_file)
                ws_yearly = wb_yearly.active
                assert ws_yearly.title == "Yearly Summary"
                
                # Should have header + 3 data rows
                assert ws_yearly.max_row == 4  # Header + 3 years
                assert ws_yearly.max_column == 7  # 7 columns
                
            except ImportError:
                pytest.skip("openpyxl not available for XLSX testing")
    
    def test_empty_positions_export_with_headers(self):
        """Test export with empty positions creates empty files with correct headers."""
        empty_positions = []
        
        with tempfile.TemporaryDirectory() as temp_dir:
            exporter = CSVExporter(temp_dir)
            
            # The exporter returns early for empty positions, so we test the individual methods
            portfolio_data = exporter._convert_positions_to_records(empty_positions)
            yearly_data = exporter._aggregate_yearly_data(empty_positions)
            
            # Both should return empty lists
            assert portfolio_data == []
            assert yearly_data == []
            
            # Test that empty data creates proper CSV files with headers
            portfolio_file = exporter._export_portfolio_history(portfolio_data)
            yearly_file = exporter._export_yearly_summary(yearly_data)
            
            assert os.path.exists(portfolio_file)
            assert os.path.exists(yearly_file)
            
            # Check portfolio history headers
            with open(portfolio_file, 'r', encoding='utf-8') as f:
                content = f.read().strip()
                lines = content.split('\n')
                assert len(lines) == 1  # Only header
                
                header = lines[0]
                expected_fields = [
                    'Date', 'Operation Description', 'Operation Quantity', 
                    'Final Quantity', 'Total Cost USD', 'Average Price USD',
                    'Total Cost BRL', 'Average Price BRL', 'Gross Profit BRL'
                ]
                for field in expected_fields:
                    assert field in header
            
            # Check yearly summary headers
            with open(yearly_file, 'r', encoding='utf-8') as f:
                content = f.read().strip()
                lines = content.split('\n')
                assert len(lines) == 1  # Only header
                
                header = lines[0]
                expected_fields = [
                    'Year', 'Final Quantity', 'Total Cost USD', 'Average Price USD',
                    'Total Cost BRL', 'Average Price BRL', 'Gross Profit BRL'
                ]
                for field in expected_fields:
                    assert field in header
    
    def test_exporter_factory_with_year_reset_data(self, multi_year_positions):
        """Test ExporterFactory creates exporters that handle year resets correctly."""
        with tempfile.TemporaryDirectory() as temp_dir:
            # Test CSV exporter from factory
            csv_exporter = ExporterFactory.create_exporter('csv', temp_dir)
            csv_exporter.export_portfolio_data(multi_year_positions)
            
            portfolio_csv = os.path.join(temp_dir, 'portfolio_history.csv')
            assert os.path.exists(portfolio_csv)
            
            # Test NoOp exporter
            noop_exporter = ExporterFactory.create_exporter('none')
            # Should not raise any errors
            noop_exporter.export_portfolio_data(multi_year_positions)
            
            # Test XLSX exporter from factory (if available)
            try:
                xlsx_exporter = ExporterFactory.create_exporter('xlsx', temp_dir)
                xlsx_exporter.export_portfolio_data(multi_year_positions)
                
                portfolio_xlsx = os.path.join(temp_dir, 'portfolio_history.xlsx')
                assert os.path.exists(portfolio_xlsx)
                
            except ImportError:
                # openpyxl not available, skip XLSX test
                pass
            
            # Test invalid format
            with pytest.raises(ValueError, match="Unsupported export format"):
                ExporterFactory.create_exporter('invalid_format')
    
    def test_export_with_single_year_no_reset(self):
        """Test export with operations in single year (no reset needed)."""
        exchange_service = MockExchangeRateService()
        calc_service = PortfolioCalculationService(exchange_service)
        
        operations = [
            VestingOperation(
                date=datetime(2023, 1, 15),
                quantity=StockQuantity(100),
                price_per_share_usd=Money(Decimal('10.00'), 'USD')
            ),
            TradeOperation(
                date=datetime(2023, 6, 20),
                quantity=StockQuantity(30),
                price_per_share_usd=Money(Decimal('15.00'), 'USD')
            ),
            VestingOperation(
                date=datetime(2023, 9, 10),
                quantity=StockQuantity(40),
                price_per_share_usd=Money(Decimal('12.00'), 'USD')
            )
        ]
        
        positions = calc_service.execute_operations(operations)
        
        with tempfile.TemporaryDirectory() as temp_dir:
            exporter = CSVExporter(temp_dir)
            exporter.export_portfolio_data(positions)
            
            # Read portfolio history
            portfolio_file = os.path.join(temp_dir, 'portfolio_history.csv')
            with open(portfolio_file, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                rows = list(reader)
            
            # All operations in same year
            years = [int(row['Date'].split('/')[-1]) for row in rows]
            assert all(year == 2023 for year in years)
            
            # Profit should accumulate (no resets)
            profits = [float(row['Gross Profit BRL']) for row in rows]
            assert profits[0] == 0.0     # Vesting
            assert profits[1] > 0.0      # Trade
            assert profits[2] == profits[1]  # Vesting (no change)
    
    def test_export_operation_descriptions_accuracy(self):
        """Test that operation descriptions are accurate in exports."""
        exchange_service = MockExchangeRateService()
        calc_service = PortfolioCalculationService(exchange_service)
        
        operations = [
            VestingOperation(
                date=datetime(2023, 1, 15),
                quantity=StockQuantity(150),
                price_per_share_usd=Money(Decimal('8.50'), 'USD')
            ),
            TradeOperation(
                date=datetime(2023, 6, 20),
                quantity=StockQuantity(75),
                price_per_share_usd=Money(Decimal('12.25'), 'USD')
            )
        ]
        
        positions = calc_service.execute_operations(operations)
        
        with tempfile.TemporaryDirectory() as temp_dir:
            exporter = CSVExporter(temp_dir)
            exporter.export_portfolio_data(positions)
            
            # Read and verify descriptions
            portfolio_file = os.path.join(temp_dir, 'portfolio_history.csv')
            with open(portfolio_file, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                rows = list(reader)
            
            # Verify operation descriptions and quantities
            assert rows[0]['Operation Description'] == 'Vesting (+150)'
            assert rows[0]['Operation Quantity'] == '150'
            assert rows[0]['Final Quantity'] == '150'
            
            assert rows[1]['Operation Description'] == 'Trade (-75)'
            assert rows[1]['Operation Quantity'] == '75'
            assert rows[1]['Final Quantity'] == '75'  # 150 - 75 = 75
    
    def test_export_with_positions_missing_operation_info(self):
        """Test export handles positions without operation information gracefully."""
        # Create positions without operation info (legacy compatibility)
        positions = [
            PortfolioPosition(
                quantity=StockQuantity(100),
                total_cost_usd=Money(Decimal('1000.00'), 'USD'),
                total_cost_brl=Money(Decimal('5000.00'), 'BRL'),
                average_price_usd=Money(Decimal('10.00'), 'USD'),
                average_price_brl=Money(Decimal('50.00'), 'BRL'),
                gross_profit_brl=Money(Decimal('200.00'), 'BRL'),
                last_updated=datetime(2023, 6, 15),
                operation_quantity=None,  # No operation info
                operation_type=None
            )
        ]
        
        with tempfile.TemporaryDirectory() as temp_dir:
            exporter = CSVExporter(temp_dir)
            exporter.export_portfolio_data(positions)
            
            # Should not crash and should handle missing info
            portfolio_file = os.path.join(temp_dir, 'portfolio_history.csv')
            with open(portfolio_file, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                rows = list(reader)
            
            assert len(rows) == 1
            assert rows[0]['Operation Description'] == 'N/A'
            assert rows[0]['Operation Quantity'] == '0'
            assert rows[0]['Final Quantity'] == '100'
