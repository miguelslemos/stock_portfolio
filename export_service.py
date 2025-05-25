import os
from typing import List
from openpyxl import Workbook
from models import PortfolioSnapshot
from operation import SellOperation
from utils import format_date


def export_to_excel(history: List[PortfolioSnapshot], yearly_summaries: List, filename: str = "portfolio_export.xlsx") -> None:
    """Export portfolio history and yearly summaries to Excel file with .xlsx extension."""
    workbook = Workbook()
    
    # Create Portfolio History sheet (detailed transactions)
    portfolio_sheet = workbook.active
    portfolio_sheet.title = 'Portfolio History'
    
    portfolio_headers = [
        'Date', 'Quantity', 'Stocks Received', 'Stock Price at Date', 'USD Quote At Date',
        'Total Cost USD', 'Average Price USD', 'Total Cost BRL', 'Average Price BRL'
    ]
    
    # Write portfolio history headers
    for col, header in enumerate(portfolio_headers, start=1):
        portfolio_sheet.cell(row=1, column=col, value=header)
    
    # Write portfolio history data
    for row, snapshot in enumerate(history, start=2):
        portfolio_sheet.cell(row=row, column=1, value=format_date(snapshot.operation.date))
        portfolio_sheet.cell(row=row, column=2, value=snapshot.total_quantity)
        
        stocks_received = snapshot.operation.quantity if not isinstance(snapshot.operation, SellOperation) else 0
        portfolio_sheet.cell(row=row, column=3, value=stocks_received)
        portfolio_sheet.cell(row=row, column=4, value=snapshot.operation.price)
        portfolio_sheet.cell(row=row, column=5, value=snapshot.usd_brl_rate)
        portfolio_sheet.cell(row=row, column=6, value=snapshot.total_cost_usd)
        portfolio_sheet.cell(row=row, column=7, value=snapshot.average_price_usd)
        portfolio_sheet.cell(row=row, column=8, value=snapshot.total_cost_brl)
        portfolio_sheet.cell(row=row, column=9, value=snapshot.average_price_brl)
    
    # Create Yearly Summary Tab
    yearly_sheet = workbook.create_sheet(title='Yearly Summary')
    
    yearly_headers = [
        'Year', 'Total Operations', 'Final Quantity', 'Total Cost USD', 
        'Average Price USD', 'Total Cost BRL', 'Average Price BRL', 'Gross Profit BRL'
    ]
    
    # Write yearly summary headers
    for col, header in enumerate(yearly_headers, start=1):
        yearly_sheet.cell(row=1, column=col, value=header)
    
    # Write yearly summary data
    for row, summary in enumerate(yearly_summaries, start=2):
        yearly_sheet.cell(row=row, column=1, value=summary.year)
        yearly_sheet.cell(row=row, column=2, value=summary.total_operations)
        yearly_sheet.cell(row=row, column=3, value=summary.final_quantity)
        yearly_sheet.cell(row=row, column=4, value=summary.total_cost_usd)
        yearly_sheet.cell(row=row, column=5, value=summary.average_price_usd)
        yearly_sheet.cell(row=row, column=6, value=summary.total_cost_brl)
        yearly_sheet.cell(row=row, column=7, value=summary.average_price_brl)
        yearly_sheet.cell(row=row, column=8, value=summary.gross_profit_brl)
    
    # Save workbook
    workbook.save(filename)
    absolute_path = os.path.abspath(filename)
    print(f"\nPortfolio data exported to: {absolute_path}")
    print(f"  - Portfolio History sheet: {len(history)} transactions")
    print(f"  - Yearly Summary sheet: {len(yearly_summaries)} years") 
