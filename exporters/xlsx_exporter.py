from abc import abstractmethod
import os
from typing import List

from openpyxl import Workbook
from exporters.file_exporter import FileExporter
from portfolio_snapshot import PortfolioSnapshot


class XLSXExporter(FileExporter):
    """Exporter for XLSX format."""
    
    def __init__(self):
        self.current_row = 1

    @staticmethod
    def get_exporter_type() -> str:
        return "xlsx"
    
    def _write_row(self, sheet, row_data):
        for col, value in enumerate(row_data, start=1):
            sheet.cell(row=self.current_row, column=col, value=value)
        self.current_row += 1

    def export(self, history: List[PortfolioSnapshot], yearly_summaries: List) -> None:
        """Export portfolio history and yearly summaries to Excel file with .xlsx extension."""
        filename: str = "portfolio_export.xlsx"
        workbook = Workbook()
        portfolio_sheet = workbook.active
        portfolio_sheet.title = 'Portfolio History'
        self._write_portfolio_history(portfolio_sheet, history)
        yearly_sheet = workbook.create_sheet(title='Yearly Summary')
        self.current_row = 1
        self._write_yearly_summary(yearly_sheet, yearly_summaries)
        workbook.save(filename)
        absolute_path = os.path.abspath(filename)
        print(f"\nPortfolio data exported to: {absolute_path}")
        print(f"  - Portfolio History sheet: {len(history)} transactions")
        print(f"  - Yearly Summary sheet: {len(yearly_summaries)} years")

